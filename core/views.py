from django.shortcuts import render, redirect
from django.views.generic import TemplateView
from django.contrib import messages
import calendar
import math
from datetime import date, datetime
from .models import CalendarEvent

def calculate_haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi / 2.0) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2.0) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

def handle_remote_checkin_notification(request, lat, lng, employee_name, check_type_str):
    if lat is None or lng is None:
        return
    active_role = request.session.get('active_role', '').upper()
    if active_role in ['SALES', 'IT', 'IT TEAM']:
        OFFICE_LAT = 27.6773098
        OFFICE_LNG = 85.3979076
        dist = calculate_haversine_distance(lat, lng, OFFICE_LAT, OFFICE_LNG)
        if dist > 50:
            from .models import SystemUserProfile, SystemNotification
            target_roles = ['HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD', 'COO']
            recipients = [u for u in SystemUserProfile.objects.all() if u.position.upper() in target_roles]
            action = "Checked In" if check_type_str == 'in' else "Checked Out"
            msg = f"{employee_name} ({active_role}) has {action} remotely from {int(dist)}m away."
            for user in recipients:
                SystemNotification.objects.create(
                    recipient=user,
                    message=msg,
                    link='/attendance/'
                )

class DashboardView(TemplateView):
    template_name = "dashboard.html"

    def get(self, request, *args, **kwargs):
        from django.utils import timezone
        now = timezone.localtime()
        import nepali_datetime
        from .nepali_calendar import get_nepali_monthdatescalendar
        
        today_np = nepali_datetime.date.today()
        year = int(request.GET.get('year', today_np.year))
        month = int(request.GET.get('month', today_np.month))

        if month < 1:
            month = 12
            year -= 1
        elif month > 12:
            month = 1
            year += 1

        # Check for initial deployment events
        if not CalendarEvent.objects.exists():
            CalendarEvent.objects.create(title="[HOLIDAY] Republic Day", date=date(2026, 5, 28), description="National Holiday")
            CalendarEvent.objects.create(title="[HOLIDAY] Dashain", date=date(2026, 10, 15), description="Festival Holiday")
            CalendarEvent.objects.create(title="CTO Review", date=date(2026, 6, 3), description="Quarterly tech audit")
            CalendarEvent.objects.create(title="Inventory Audit", date=date(2026, 6, 10), description="HQ Stock Verification")
            CalendarEvent.objects.create(title="Payroll Sync", date=date(2026, 6, 19), description="June salaries sign-off")
            CalendarEvent.objects.create(title="System Patch", date=date(2026, 6, 25), description="Core ERP upgrade")

        weeks = get_nepali_monthdatescalendar(year, month)

        start_date = weeks[0][0]
        end_date = weeks[-1][-1]
        events = CalendarEvent.objects.filter(date__range=(start_date, end_date))

        events_by_date = {}
        for event in events:
            events_by_date.setdefault(event.date, []).append(event)

        formatted_weeks = []
        for week in weeks:
            formatted_week = []
            for day in week:
                day_np = nepali_datetime.date.from_datetime_date(day)
                is_current_month = (day_np.month == month)
                is_today = (day == now.date())
                day_events = events_by_date.get(day, [])
                is_holiday = any('[HOLIDAY]' in e.title for e in day_events)
                formatted_week.append({
                    'date': day,
                    'day_num': day_np.day, # Use Nepali day for rendering
                    'is_current_month': is_current_month,
                    'is_today': is_today,
                    'events': day_events,
                    'is_holiday': is_holiday
                })
            formatted_weeks.append(formatted_week)

        prev_month = month - 1
        prev_year = year
        if prev_month < 1:
            prev_month = 12
            prev_year -= 1

        next_month = month + 1
        next_year = year
        if next_month > 12:
            next_month = 1
            next_year += 1

        # Nepali month names
        np_months = ["Baisakh", "Jestha", "Ashadh", "Shrawan", "Bhadra", "Ashwin", "Kartik", "Mangsir", "Poush", "Magh", "Falgun", "Chaitra"]
        month_name = np_months[month - 1]
        
        # Calculate Dashboard Activity (Leaves & Missing Attendance)
        from .models import LeaveRequest, AttendanceRecord, SystemUserProfile
        today = now.date()
        
        on_leave_today = LeaveRequest.objects.filter(
            status='Approved',
            start_date__lte=today,
            end_date__gte=today
        )
        
        attended_records = AttendanceRecord.objects.filter(
            date=today,
            check_in_time__isnull=False,
            latitude__isnull=False
        )
        on_leave_names = on_leave_today.values_list('employee_name', flat=True)
        
        all_users = SystemUserProfile.objects.exclude(full_name__in=list(on_leave_names))
        attendance_status = []
        for user in all_users:
            record = attended_records.filter(employee_name=user.full_name).first()
            if record:
                status = 'checked_out' if record.check_out_time else 'checked_in'
            else:
                status = 'not_checked_in'
                
            attendance_status.append({
                'user': user,
                'status': status
            })

        active_role = request.session.get('active_role', '')
        is_remote_allowed = active_role.upper() in ['SALES', 'IT', 'IT TEAM']

        logged_in_name = request.session.get('logged_in_name', 'John Doe')
        today_record = AttendanceRecord.objects.filter(employee_name=logged_in_name, date=today).first()
        is_checked_in = False
        is_checked_out = False
        if today_record:
            if today_record.check_in_time:
                is_checked_in = True
            if today_record.check_out_time:
                is_checked_out = True

        context = {
            'weeks': formatted_weeks,
            'year': year,
            'month': month,
            'month_name': month_name,
            'prev_year': prev_year,
            'prev_month': prev_month,
            'next_year': next_year,
            'next_month': next_month,
            'on_leave_today': on_leave_today,
            'attendance_status': attendance_status,
            'is_remote_allowed': is_remote_allowed,
            'is_checked_in': is_checked_in,
            'is_checked_out': is_checked_out,
            'today_record': today_record,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        if action == 'delete_event':
            active_role = request.session.get('active_role', '')
            if active_role.upper() in ['CEO', 'COO', 'HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD', 'CITO', 'CHAIRMAN', 'OPERATION', 'ADMIN']:
                event_id = request.POST.get('event_id')
                if event_id:
                    CalendarEvent.objects.filter(id=event_id).delete()
            year = request.POST.get('year', request.GET.get('year', ''))
            month = request.POST.get('month', request.GET.get('month', ''))
            url = '/'
            if year and month:
                url += f'?year={year}&month={month}'
            return redirect(url)

        title = request.POST.get('title')
        description = request.POST.get('description', '')
        date_str = request.POST.get('date')
        is_holiday = request.POST.get('is_holiday') == 'on'
        
        if title and is_holiday:
            title = f"[HOLIDAY] {title}"

        if title and date_str:
            try:
                import nepali_datetime
                # Parse the incoming Nepali date (YYYY-MM-DD)
                year_part, month_part, day_part = map(int, date_str.split('-'))
                np_date = nepali_datetime.date(year_part, month_part, day_part)
                # Convert to standard Gregorian date for database storage
                event_date = np_date.to_datetime_date()
                
                CalendarEvent.objects.create(
                    title=title,
                    description=description,
                    date=event_date
                )
            except (ValueError, TypeError):
                pass

        if action == 'check_in_out':
            from .models import AttendanceRecord
            from datetime import date, datetime, time
            from django.utils import timezone
            logged_in_name = request.session.get('logged_in_name', 'John Doe')
            lat_str = request.POST.get('latitude')
            lng_str = request.POST.get('longitude')
            now = timezone.localtime()
            today = now.date()
            current_time = now.time()

            lat = None
            lng = None
            if lat_str and lng_str:
                try:
                    lat = float(lat_str)
                    lng = float(lng_str)
                except ValueError:
                    pass

            cutoff_time = time(9, 30, 0)
            status = 'Present'
            if current_time > cutoff_time:
                status = 'Late'

            record, created = AttendanceRecord.objects.get_or_create(
                employee_name=logged_in_name,
                date=today,
                defaults={
                    'check_in_time': current_time,
                    'status': status,
                    'latitude': lat,
                    'longitude': lng
                }
            )

            is_new_checkout = False
            check_type = request.POST.get('check_type')
            if check_type == 'in':
                pass # Already checked in
            elif not created and not record.check_out_time:
                record.check_out_time = current_time
                record.save()
                is_new_checkout = True
                
            if created or is_new_checkout:
                c_type = 'in' if created else 'out'
                handle_remote_checkin_notification(request, lat, lng, logged_in_name, c_type)

            from django.http import JsonResponse
            return JsonResponse({
                'success': True,
                'checked_in': record.check_in_time is not None,
                'checked_out': record.check_out_time is not None,
                'check_in_time': record.check_in_time.strftime('%I:%M %p') if record.check_in_time else '',
                'check_out_time': record.check_out_time.strftime('%I:%M %p') if record.check_out_time else '',
                'status': record.status
            })

        year = request.POST.get('year', '')
        month = request.POST.get('month', '')
        url = '/'
        if year and month:
            url += f'?year={year}&month={month}'
        return redirect(url)

from .models import LeaveRequest, AttendanceRecord
import calendar
from datetime import date, datetime, time

class AttendanceCheckinView(TemplateView):
    template_name = "attendance_checkin.html"

    def get(self, request, *args, **kwargs):
        import nepali_datetime
        from .nepali_calendar import get_nepali_monthdatescalendar
        
        # Pre-populate simulated attendance history for the user for the current month if empty
        logged_in_name = request.session.get('logged_in_name', 'John Doe')
        active_role = request.session.get('active_role', '')
        
        exec_roles = ['CEO', 'COO', 'HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD']
        can_view_all = active_role.upper() in exec_roles
        
        view_user_name = request.GET.get('view_user_name', logged_in_name)
        if not can_view_all:
            view_user_name = logged_in_name
            
        from .models import SystemUserProfile
        all_users = []
        if can_view_all:
            all_users = SystemUserProfile.objects.values('full_name', 'uid')

        from django.utils import timezone
        today = timezone.localtime().date()
        today_np = nepali_datetime.date.today()
        year = int(request.GET.get('year', today_np.year))
        month = int(request.GET.get('month', today_np.month))

        if month < 1:
            month = 12
            year -= 1
        elif month > 12:
            month = 1
            year += 1

        if not AttendanceRecord.objects.filter(employee_name=logged_in_name).exists():
            # Populate some days in June 2026 (or current month)
            for d in range(1, 19):
                day_date = date(2026, 6, d)
                weekday = day_date.weekday()
                if weekday != 5: # Sunday - Friday (Saturday is weekend)
                    # Randomize a bit: Present or Late
                    if d == 12:
                        AttendanceRecord.objects.create(
                            employee_name=logged_in_name,
                            date=day_date,
                            check_in_time=time(9, 45, 0),
                            check_out_time=time(17, 30, 0),
                            status="Late",
                            latitude=27.6773098,
                            longitude=85.3979076
                        )
                    elif d == 15:
                        AttendanceRecord.objects.create(
                            employee_name=logged_in_name,
                            date=day_date,
                            status="Leave"
                        )
                    else:
                        AttendanceRecord.objects.create(
                            employee_name=logged_in_name,
                            date=day_date,
                            check_in_time=time(8, 55, 0),
                            check_out_time=time(17, 0, 0),
                            status="Present",
                            latitude=27.6773098,
                            longitude=85.3979076
                        )

        # Generate month grid
        weeks = get_nepali_monthdatescalendar(year, month)
        
        start_date = weeks[0][0]
        end_date = weeks[-1][-1]

        # Fetch records for selected month
        records = AttendanceRecord.objects.filter(
            employee_name=view_user_name,
            date__range=(start_date, end_date)
        )
        records_by_date = {r.date: r for r in records}

        # Calculate statistics
        # Only count for days strictly in the selected nepali month
        monthly_records = []
        for w in weeks:
            for d in w:
                day_np = nepali_datetime.date.from_datetime_date(d)
                if day_np.month == month and d in records_by_date:
                    monthly_records.append(records_by_date[d])

        total_present = sum(1 for r in monthly_records if r.status == 'Present')
        total_late = sum(1 for r in monthly_records if r.status == 'Late')
        total_leave = sum(1 for r in monthly_records if r.status == 'Leave')
        total_absent = sum(1 for r in monthly_records if r.status == 'Absent')

        # Build grid list
        formatted_weeks = []
        for week in weeks:
            current_week = []
            for col, day_date in enumerate(week):
                day_np = nepali_datetime.date.from_datetime_date(day_date)
                is_current_month = (day_np.month == month)
                
                if not is_current_month:
                    current_week.append({
                        'day': '',
                        'record': None,
                        'is_today': False,
                        'is_weekend': False
                    })
                else:
                    is_weekend = col == 6 # Saturday=6 for start Sunday? Wait, get_nepali_monthdatescalendar returns Sunday first usually.
                    # Let's check: weekday() usually has 0=Monday, 6=Sunday. In our grid Saturday is weekend.
                    is_weekend = (day_date.weekday() == 5)
                    record = records_by_date.get(day_date)
                    is_today = (day_date == today)

                    current_week.append({
                        'day': day_np.day,
                        'record': record,
                        'is_today': is_today,
                        'is_weekend': is_weekend
                    })
            formatted_weeks.append(current_week)

        # Date Nav variables
        np_months = ["Baisakh", "Jestha", "Ashadh", "Shrawan", "Bhadra", "Ashwin", "Kartik", "Mangsir", "Poush", "Magh", "Falgun", "Chaitra"]
        month_name = np_months[month - 1]
        
        prev_month = month - 1
        prev_year = year
        if prev_month < 1:
            prev_month = 12
            prev_year -= 1

        next_month = month + 1
        next_year = year
        if next_month > 12:
            next_month = 1
            next_year += 1

        active_role = request.session.get('active_role', '')
        is_remote_allowed = active_role.upper() in ['SALES', 'IT', 'IT TEAM']

        # Determine check-in/out status for today
        today_record = AttendanceRecord.objects.filter(employee_name=logged_in_name, date=today).first()
        is_checked_in = False
        is_checked_out = False
        if today_record:
            if today_record.check_in_time:
                is_checked_in = True
            if today_record.check_out_time:
                is_checked_out = True

        context = {
            'weeks': formatted_weeks,
            'year': year,
            'month': month,
            'month_name': month_name,
            'prev_year': prev_year,
            'prev_month': prev_month,
            'next_year': next_year,
            'next_month': next_month,
            'total_present': total_present,
            'total_late': total_late,
            'total_leave': total_leave,
            'total_absent': total_absent,
            'today': today,
            'is_remote_allowed': is_remote_allowed,
            'is_checked_in': is_checked_in,
            'is_checked_out': is_checked_out,
            'today_record': today_record,
            'all_users': all_users,
            'view_user_name': view_user_name,
            'can_view_all': can_view_all,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        from django.utils import timezone
        from datetime import time, datetime
        
        action = request.POST.get('action')
        if action == 'missed_attendance':
            logged_in_name = request.session.get('logged_in_name', 'John Doe')
            from .models import SystemUserProfile
            system_user = SystemUserProfile.objects.filter(full_name=logged_in_name).first()
            if not system_user:
                from django.contrib import messages
                messages.error(request, 'User profile not found in system.')
                return redirect('attendance_checkin')
                
            start_date_str = request.POST.get('start_date')
            end_date_str = request.POST.get('end_date')
            check_in_str = request.POST.get('check_in_time')
            check_out_str = request.POST.get('check_out_time')
            reason = request.POST.get('reason')
            
            try:
                ma_start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                ma_end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else ma_start_date
                in_time = datetime.strptime(check_in_str, '%H:%M').time()
                out_time = datetime.strptime(check_out_str, '%H:%M').time()
                
                from .models import MissedAttendanceRequest
                MissedAttendanceRequest.objects.create(
                    employee=system_user,
                    start_date=ma_start_date,
                    end_date=ma_end_date,
                    check_in_time=in_time,
                    check_out_time=out_time,
                    reason=reason
                )
                from django.contrib import messages
                messages.success(request, 'Missed attendance request submitted successfully.')
                return redirect('attendance_checkin')
            except Exception as e:
                from django.contrib import messages
                messages.error(request, f'Error submitting request: {str(e)}')
                return redirect('attendance_checkin')
        
        logged_in_name = request.session.get('logged_in_name', 'John Doe')
        lat_str = request.POST.get('latitude')
        lng_str = request.POST.get('longitude')
        
        now = timezone.localtime()
        today = now.date()
        current_time = now.time()


        lat = None
        lng = None
        if lat_str and lng_str:
            try:
                lat = float(lat_str)
                lng = float(lng_str)
            except ValueError:
                pass

        cutoff_time = time(9, 30, 0)
        status = 'Present'
        if current_time > cutoff_time:
            status = 'Late'

        record, created = AttendanceRecord.objects.get_or_create(
            employee_name=logged_in_name,
            date=today,
            defaults={
                'check_in_time': current_time,
                'status': status,
                'latitude': lat,
                'longitude': lng
            }
        )

        is_new_checkout = False
        check_type = request.POST.get('check_type')
        if check_type == 'in':
            pass
        elif not created and not record.check_out_time:
            record.check_out_time = current_time
            record.save()
            is_new_checkout = True

        if created or is_new_checkout:
            c_type = 'in' if created else 'out'
            handle_remote_checkin_notification(request, lat, lng, logged_in_name, c_type)

        from django.http import JsonResponse
        return JsonResponse({
            'success': True,
            'checked_in': record.check_in_time is not None,
            'checked_out': record.check_out_time is not None,
            'check_in_time': record.check_in_time.strftime('%I:%M %p') if record.check_in_time else '',
            'check_out_time': record.check_out_time.strftime('%I:%M %p') if record.check_out_time else '',
            'status': record.status
        })

class InventoryListView(TemplateView):
    template_name = "inventory_list.html"
    
    def get(self, request, *args, **kwargs):
        from .models import InventoryLedger, AccountsReceivable
        raw_inventory = InventoryLedger.objects.all()
        
        aggregated_data = {}
        for item in raw_inventory:
            loc = item.node_id.name if item.node_id else 'Unknown'
            sku = item.sku_id
            key = (loc, sku)
            
            if key not in aggregated_data:
                aggregated_data[key] = {
                    'location': loc,
                    'sku_id': sku,
                    'item_desc': item.product_name,
                    'category': item.category,
                    'uom': item.uom,
                    'fresh_qty': 0, 'fresh_amt': 0.0,
                    'dam_qty': 0, 'dam_amt': 0.0,
                    'refurb_qty': 0, 'refurb_amt': 0.0,
                    'rep_qty': 0, 'rep_amt': 0.0,
                    'tot_qty': 0, 'tot_amt': 0.0,
                }
            
            status = item.locator_bin_status
            qty = item.quantity_on_hand
            amt = float(qty * item.price)
            
            if status == 'Fresh':
                aggregated_data[key]['fresh_qty'] += qty
                aggregated_data[key]['fresh_amt'] += amt
            elif status == 'Damaged':
                aggregated_data[key]['dam_qty'] += qty
                aggregated_data[key]['dam_amt'] += amt
            elif status == 'Refurbished':
                aggregated_data[key]['refurb_qty'] += qty
                aggregated_data[key]['refurb_amt'] += amt
            elif status == 'Repairable':
                aggregated_data[key]['rep_qty'] += qty
                aggregated_data[key]['rep_amt'] += amt
                
            aggregated_data[key]['tot_qty'] += qty
            aggregated_data[key]['tot_amt'] += amt

        inventory = list(aggregated_data.values())
        accounts = AccountsReceivable.objects.all().order_by('-created_at')[:5]
        return render(request, self.template_name, {
            'inventory': inventory,
            'accounts': accounts
        })
        
    def post(self, request, *args, **kwargs):
        from .models import InventoryLedger, OrgNode
        action = request.POST.get('action')
        
        if action == 'delete':
            active_role = request.session.get('active_role', '')
            system_user = getattr(request, 'system_user', None)
            is_ceo = (active_role == 'CEO') or (system_user and system_user.position == 'CEO')
            
            if is_ceo:
                sku_id = request.POST.get('sku_id')
                node_name = request.POST.get('node_name')
                if sku_id and node_name:
                    InventoryLedger.objects.filter(sku_id=sku_id, node_id__name=node_name).delete()
                    from django.contrib import messages
                    messages.success(request, f'Stock details for SKU {sku_id} at {node_name} deleted successfully.')
                else:
                    from django.contrib import messages
                    messages.error(request, 'Missing SKU ID or Location for deletion.')
            else:
                from django.contrib import messages
                messages.error(request, 'Access Denied: Only the CEO can delete stock tracking details.')
            return redirect('inventory_list')
        
        if action == 'add_stock':
            sku_id = request.POST.get('sku_id')
            product_name = request.POST.get('product_name')
            category = request.POST.get('category', 'Electronics')
            uom = request.POST.get('uom', 'Pcs')
            price = request.POST.get('price')
            
            location_name = request.POST.get('location', 'HQ-NEPAL').upper()
            
            fresh_qty = int(request.POST.get('fresh_qty') or 0)
            dam_qty = int(request.POST.get('dam_qty') or 0)
            refurb_qty = int(request.POST.get('refurb_qty') or 0)
            rep_qty = int(request.POST.get('rep_qty') or 0)
            
            product_image = request.FILES.get('product_image')
            
            node, _ = OrgNode.objects.get_or_create(name=location_name, defaults={'node_type': 'Warehouse'})
            
            if fresh_qty == 0 and dam_qty == 0 and refurb_qty == 0 and rep_qty == 0:
                fresh_qty = 1
                
            if fresh_qty > 0:
                InventoryLedger.objects.create(
                    sku_id=sku_id, product_name=product_name, category=category, uom=uom,
                    price=price, quantity_on_hand=fresh_qty, locator_bin_status='Fresh',
                    node_id=node, product_image=product_image
                )
            if dam_qty > 0:
                InventoryLedger.objects.create(
                    sku_id=sku_id, product_name=product_name, category=category, uom=uom,
                    price=price, quantity_on_hand=dam_qty, locator_bin_status='Damaged',
                    node_id=node, product_image=product_image
                )
            if refurb_qty > 0:
                InventoryLedger.objects.create(
                    sku_id=sku_id, product_name=product_name, category=category, uom=uom,
                    price=price, quantity_on_hand=refurb_qty, locator_bin_status='Refurbished',
                    node_id=node, product_image=product_image
                )
            if rep_qty > 0:
                InventoryLedger.objects.create(
                    sku_id=sku_id, product_name=product_name, category=category, uom=uom,
                    price=price, quantity_on_hand=rep_qty, locator_bin_status='Repairable',
                    node_id=node, product_image=product_image
                )
                
            from django.contrib import messages
            total_qty = fresh_qty + dam_qty + refurb_qty + rep_qty
            messages.success(request, f"Successfully added {total_qty} units of {product_name} to {location_name} inventory.")
            
        elif action == 'bulk_import':
            excel_file = request.FILES.get('excel_file')
            if excel_file:
                import csv
                import io
                import openpyxl
                from django.contrib import messages
                
                filename = excel_file.name.lower()
                rows = []
                
                try:
                    if filename.endswith('.csv'):
                        decoded_file = excel_file.read().decode('utf-8-sig')
                        io_string = io.StringIO(decoded_file)
                        reader = csv.reader(io_string)
                        rows = list(reader)[1:] # Skip header
                    else:
                        wb = openpyxl.load_workbook(excel_file, data_only=True)
                        ws = wb.active
                        rows = list(ws.iter_rows(min_row=2, values_only=True))

                    count = 0
                    for row in rows:
                        row = list(row) + [None] * max(0, 10 - len(row))
                        
                        if row[0] and row[1] and row[4] and row[5] is not None:
                            sku_id = str(row[0]).strip()
                            product_name = str(row[1]).strip()
                            category = str(row[2]).strip() if row[2] else 'Electronics'
                            uom = str(row[3]).strip() if row[3] else 'Pcs'
                            location_name = str(row[4]).strip().upper()
                            
                            try:
                                price = float(str(row[5]).replace(',', '').strip())
                            except ValueError:
                                price = 0.0

                            def parse_qty(val):
                                try:
                                    return int(float(str(val).replace(',', '').strip())) if val and str(val).strip() else 0
                                except ValueError:
                                    return 0

                            fresh_qty = parse_qty(row[6])
                            dam_qty = parse_qty(row[7])
                            refurb_qty = parse_qty(row[8])
                            rep_qty = parse_qty(row[9])

                            node, _ = OrgNode.objects.get_or_create(name=location_name, defaults={'node_type': 'Warehouse'})

                            if fresh_qty > 0:
                                InventoryLedger.objects.create(sku_id=sku_id, product_name=product_name, category=category, uom=uom, price=price, quantity_on_hand=fresh_qty, locator_bin_status='Fresh', node_id=node)
                            if dam_qty > 0:
                                InventoryLedger.objects.create(sku_id=sku_id, product_name=product_name, category=category, uom=uom, price=price, quantity_on_hand=dam_qty, locator_bin_status='Damaged', node_id=node)
                            if refurb_qty > 0:
                                InventoryLedger.objects.create(sku_id=sku_id, product_name=product_name, category=category, uom=uom, price=price, quantity_on_hand=refurb_qty, locator_bin_status='Refurbished', node_id=node)
                            if rep_qty > 0:
                                InventoryLedger.objects.create(sku_id=sku_id, product_name=product_name, category=category, uom=uom, price=price, quantity_on_hand=rep_qty, locator_bin_status='Repairable', node_id=node)

                            count += 1

                    messages.success(request, f'Successfully imported {count} stock records.')
                except Exception as e:
                    messages.error(request, f'Failed to import file: {str(e)}')
            else:
                from django.contrib import messages
                messages.error(request, 'No file provided for bulk import.')
        return redirect('inventory_list')

from .models import SystemUserProfile

class UserCreateView(TemplateView):
    template_name = "user_create.html"

    def get(self, request, *args, **kwargs):
        # Auto pre-populate the 4 roles if empty
        if not SystemUserProfile.objects.exists():
            SystemUserProfile.objects.create(
                full_name="Rojil Thapa",
                position="CEO",
                profile_image="https://lh3.googleusercontent.com/aida-public/AB6AXuAxRwBqcBsIwhl8UfYuUmsQHvIz1o1r9OuPPywbOtlGXWlCR_nVlogNz1dWjHxbbR8D4JRp6xcnJNC-9wb-yTw4b0zCNhB4X2AYyTeqO9RSpOO2DlQq6-bWMWXr3BYnulb9WcLu3g_1_7WBjKHzPfS1wtbtw9TNLG4uNGiafbzvLA8kbIQtntZeSPb1stxr6gciJC4rA8CwhZUk9JLERRW8332qacALyXCTp9poQmehlr-eWspkIhqdUrJlvlr45Z3nBmWrNROkitA",
                node="HQ-NEPAL",
                pan_number="601122334",
                citizenship_number="27-01-72-99881",
                contact_number="+977-9851011223",
                email="rojil@nightvision.com.np",
                address="Kathmandu, Nepal",
                global_telemetry=True,
                ledger_override=True,
                api_key_gen=True,
                audit_log_purge=True,
                hr_matrix_view=True,
                fleet_command=True,
                can_access_profiles=True,
                can_access_lead_pipeline=True,
                can_access_inventory=True,
                can_access_accounts_receivable=True,
                can_access_pos=True,
                can_access_procurement=True
            )
            SystemUserProfile.objects.create(
                full_name="Ajay Bakhunchhe",
                position="CITO",
                profile_image="https://lh3.googleusercontent.com/aida-public/AB6AXuBs4lExNu2CzthZ9OvocXvEmbE3wXd1duo2qcQxQWFi1H3SjSgQPj1DiBb_Ct1nt7uQQsYoaMn2hteSpaPGXY83PKQXkfY-MHerms63L1zc8b71EJDM68-NCxLo33P2OaJ-MZBD_UhONo8mArPudj7a4QCtUIDrwjg2tFJothkcyey2MH2npu5ak2u5DowyJ-Qopkd6HlZ02nU7VZc1NzQCvruFJSoJyKS617KiHgPaiM20GWpTzw4wjI6Ry7pWM_1XE8ojwjonNyU",
                node="HQ-NEPAL",
                pan_number="602233445",
                citizenship_number="27-01-74-88772",
                contact_number="+977-9851022334",
                email="ajay@nightvision.com.np",
                address="Lalitpur, Nepal",
                global_telemetry=True,
                ledger_override=False,
                api_key_gen=True,
                audit_log_purge=True,
                hr_matrix_view=False,
                fleet_command=True,
                can_access_profiles=True,
                can_access_lead_pipeline=True,
                can_access_inventory=True,
                can_access_accounts_receivable=True,
                can_access_pos=True,
                can_access_procurement=True
            )
            SystemUserProfile.objects.create(
                full_name="Manjil Deuja",
                position="COO",
                profile_image="https://lh3.googleusercontent.com/aida-public/AB6AXuDWIAfodMQAXvKiYJy0D3dKaJM90Y83Tr-GC7hDuyN3qCB0yTqKnLuDstAPHjM0wwVSGazyg8CcsFcQ1hHACzcemG4TZ9uolATIn0Dn9j0Ua4hPmEwCljlUW6MQcop18cIyvYknO0jevfZHPeoS2Ax32DYkBi8-fdaKBtjVZq8FbvG2pbYcgIis8PDZI4LkiQkC01y56-3BhHabvA8mA5NoIElPYdRJs8lgQjNJg9CsRoDuKtYIesv8UOyJUfTfzFskb5ia0Z_dMHU",
                node="HQ-NEPAL",
                pan_number="603344556",
                citizenship_number="27-01-76-77663",
                contact_number="+977-9851033445",
                email="manjil@nightvision.com.np",
                address="Bhaktapur, Nepal",
                global_telemetry=True,
                ledger_override=True,
                api_key_gen=False,
                audit_log_purge=False,
                hr_matrix_view=True,
                fleet_command=True,
                can_access_profiles=True,
                can_access_lead_pipeline=True,
                can_access_inventory=True,
                can_access_accounts_receivable=True,
                can_access_pos=True,
                can_access_procurement=True
            )
            SystemUserProfile.objects.create(
                full_name="Hemanta Mahato",
                position="HR and Operation Head",
                profile_image="https://lh3.googleusercontent.com/aida-public/AB6AXuAVDI2iFpujCL_v3DHDuqdNMFY7p4wtgmzwORSUfv0PapsXSmFDOi4AUHB2lVdjX-JUWZaSbslLRCw-6ClmiJyxgMxcPwSfu2dLNPoCEiN2ouO5wsu9qOwgI5RljBwvQvtFoEjWRfrh_zJixxJaDH3D2AB4fEOi9_cvcfq4ybvBoRjgIBwnjIkWsKEr-NAalgRTBpnbgNgWocmp8qJKR5YrjjUHAhmcclkDIT4pETOpZalJP8DC2UYsqoi0rLPSjkkhA6squGCp88k",
                node="HQ-NEPAL",
                pan_number="604455667",
                citizenship_number="27-01-78-66554",
                contact_number="+977-9851044556",
                email="hemanta@nightvision.com.np",
                address="Kirtipur, Nepal",
                global_telemetry=True,
                ledger_override=False,
                api_key_gen=False,
                audit_log_purge=False,
                hr_matrix_view=True,
                fleet_command=True,
                can_access_profiles=True,
                can_access_lead_pipeline=True,
                can_access_inventory=True,
                can_access_accounts_receivable=True,
                can_access_pos=True,
                can_access_procurement=True
            )

        logged_in_name = request.session.get('logged_in_name', '')
        from django.db.models import Case, When, Value, IntegerField
        users = SystemUserProfile.objects.all().annotate(
            is_current=Case(
                When(full_name=logged_in_name, then=Value(0)),
                default=Value(1),
                output_field=IntegerField()
            )
        ).order_by('is_current', 'created_at')
        selected_uid = request.GET.get('uid')
        
        # Restriction for Staff
        active_role = request.session.get('active_role')
        if active_role == 'Staff':
            logged_in_uid = request.session.get('logged_in_uid')
            if selected_uid and selected_uid != logged_in_uid:
                return redirect(f'/users/create/?uid={logged_in_uid}')
            users = users.filter(uid=logged_in_uid)
            selected_uid = logged_in_uid

        selected_user = None

        if selected_uid:
            selected_user = users.filter(uid=selected_uid).first()
        if not selected_user:
            selected_user = users.first()

        can_view_documents = False
        documents = []
        if selected_user:
            logged_in_uid = request.session.get('logged_in_uid')
            active_role = request.session.get('active_role', '')
            if active_role.upper() in ['CEO', 'COO', 'HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD'] or selected_user.uid == logged_in_uid:
                can_view_documents = True
                documents = selected_user.documents.all().order_by('-uploaded_at')

        # Get distinct positions dynamically
        distinct_positions = list(SystemUserProfile.objects.values_list('position', flat=True).distinct())
        default_positions = ['CEO', 'CITO', 'IT', 'COO', 'HR and Operation Head', 'Admin', 'CMO', 'CSO', 'CTO', 'Manager', 'Logistics Manager', 'Staff']
        for p in default_positions:
            if p not in distinct_positions:
                distinct_positions.append(p)
        distinct_positions.sort()

        context = {
            'users': users,
            'selected_user': selected_user,
            'can_view_documents': can_view_documents,
            'documents': documents,
            'distinct_positions': distinct_positions,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        uid = request.POST.get('uid')

        active_role = request.session.get('active_role')
        if active_role == 'Staff':
            if action != 'update' or uid != request.session.get('logged_in_uid'):
                return redirect(f"/users/create/?uid={request.session.get('logged_in_uid')}&error=unauthorized")

        if action == 'create':
            full_name = request.POST.get('full_name')
            position = request.POST.get('position') or 'Staff'
            node = request.POST.get('node', 'HQ-NEPAL')
            password = request.POST.get('password', 'Admin')
            profile_image_input = request.POST.get('profile_image')
            
            pan_number = request.POST.get('pan_number', '')
            citizenship_number = request.POST.get('citizenship_number', '')
            contact_number = request.POST.get('contact_number', '')
            emergency_contact_number = request.POST.get('emergency_contact_number', '')
            email = request.POST.get('email', '')
            address = request.POST.get('address', '')
            
            gender = request.POST.get('gender', 'Male')
            educational_qualifications = request.POST.get('educational_qualifications')
            languages_known = request.POST.get('languages_known')
            proficiency_level = request.POST.get('proficiency_level')

            # Read toggles
            global_telemetry = request.POST.get('global_telemetry') == 'true'
            ledger_override = request.POST.get('ledger_override') == 'true'
            api_key_gen = request.POST.get('api_key_gen') == 'true'
            audit_log_purge = request.POST.get('audit_log_purge') == 'true'
            hr_matrix_view = request.POST.get('hr_matrix_view') == 'true'
            fleet_command = request.POST.get('fleet_command') == 'true'

            # File Upload Handling
            profile_image = None
            profile_image_file = request.FILES.get('profile_image_file')
            if profile_image_file:
                from django.core.files.storage import FileSystemStorage
                fs = FileSystemStorage()
                filename = fs.save('avatars/' + profile_image_file.name, profile_image_file)
                profile_image = fs.url(filename)
            elif profile_image_input:
                profile_image = profile_image_input
            else:
                avatars = [
                    "https://images.unsplash.com/photo-1534528741775-53994a69daeb?auto=format&fit=crop&w=150&q=80",
                    "https://images.unsplash.com/photo-1507003211169-0a1dd7228f2d?auto=format&fit=crop&w=150&q=80",
                    "https://images.unsplash.com/photo-1500648767791-00dcc994a43e?auto=format&fit=crop&w=150&q=80",
                    "https://images.unsplash.com/photo-1494790108377-be9c29b29330?auto=format&fit=crop&w=150&q=80"
                ]
                import random
                profile_image = random.choice(avatars)

            # Assign all page accesses by default to high-level roles
            is_senior = position.upper() in ['CEO', 'CITO', 'COO', 'HR AND OPERATION HEAD', 'ADMIN', 'SYSTEM ADMIN', 'HR', 'OPERATION HEAD', 'CHAIRMAN', 'OPERATION']

            user = SystemUserProfile.objects.create(
                full_name=full_name,
                position=position,
                node=node,
                password=password,
                profile_image=profile_image,
                pan_number=pan_number,
                citizenship_number=citizenship_number,
                contact_number=contact_number,
                emergency_contact_number=emergency_contact_number,
                email=email,
                address=address,
                gender=gender,
                educational_qualifications=educational_qualifications,
                languages_known=languages_known,
                proficiency_level=proficiency_level,
                global_telemetry=global_telemetry,
                ledger_override=ledger_override,
                api_key_gen=api_key_gen,
                audit_log_purge=audit_log_purge,
                hr_matrix_view=hr_matrix_view,
                fleet_command=fleet_command,
                can_access_dashboard=request.POST.get('can_access_dashboard') == 'true' if 'can_access_dashboard' in request.POST else True,
                can_access_notice_board=request.POST.get('can_access_notice_board') == 'true' if 'can_access_notice_board' in request.POST else True,
                can_access_time_attendance=request.POST.get('can_access_time_attendance') == 'true' if 'can_access_time_attendance' in request.POST else True,
                can_access_leave=request.POST.get('can_access_leave') == 'true' if 'can_access_leave' in request.POST else True,
                can_access_profiles=request.POST.get('can_access_profiles') == 'true' if 'can_access_profiles' in request.POST else is_senior,
                can_access_lead_pipeline=request.POST.get('can_access_lead_pipeline') == 'true' if 'can_access_lead_pipeline' in request.POST else is_senior,
                can_access_inventory=request.POST.get('can_access_inventory') == 'true' if 'can_access_inventory' in request.POST else is_senior,
                can_access_accounts_receivable=request.POST.get('can_access_accounts_receivable') == 'true' if 'can_access_accounts_receivable' in request.POST else is_senior,
                can_access_pos=request.POST.get('can_access_pos') == 'true' if 'can_access_pos' in request.POST else is_senior,
                can_access_procurement=request.POST.get('can_access_procurement') == 'true' if 'can_access_procurement' in request.POST else is_senior,
            )
            
            if user.gender == 'Female':
                from .models import CalendarEvent, LeaveRequest
                female_holidays = CalendarEvent.objects.filter(is_female_holiday=True)
                for fh in female_holidays:
                    LeaveRequest.objects.get_or_create(
                        employee_name=user.full_name,
                        leave_type='Casual',
                        start_date=fh.date,
                        end_date=fh.date,
                        defaults={
                            'reason': f'Female Holiday: {fh.title}',
                            'status': 'Approved'
                        }
                    )
            
            return redirect(f'/users/create/?uid={user.uid}')

        elif action == 'update':
            import os
            from django.conf import settings
            from django.core.files.storage import FileSystemStorage
            uid = request.POST.get('uid')
            user = SystemUserProfile.objects.filter(uid=uid).first()
            if user:
                old_name = user.full_name
                user.full_name = request.POST.get('full_name', user.full_name)
                
                active_role_upper = request.session.get('active_role', '').upper()
                
                # Update UID safely
                if active_role_upper in ['CEO', 'COO']:
                    new_uid = request.POST.get('new_uid', '').strip()
                    if new_uid and new_uid != user.uid:
                        # Ensure no duplicate
                        if not SystemUserProfile.objects.filter(uid=new_uid).exists():
                            user.uid = new_uid
                            # If self-editing, update session
                            if request.session.get('logged_in_uid') == uid:
                                request.session['logged_in_uid'] = new_uid

                if active_role_upper in ['CEO', 'COO', 'HR AND OPERATION HEAD']:
                    user.position = request.POST.get('position', user.position)
                    
                user.node = request.POST.get('node', user.node)
                
                password = request.POST.get('password', '').strip()
                if password:
                    user.password = password
                    
                profile_image_file = request.FILES.get('profile_image_file')
                if profile_image_file:
                    fs = FileSystemStorage()
                    filename = fs.save('profiles/' + profile_image_file.name, profile_image_file)
                    user.profile_image = fs.url(filename)
                else:
                    user.profile_image = request.POST.get('profile_image', user.profile_image)
                
                user.pan_number = request.POST.get('pan_number', user.pan_number)
                user.citizenship_number = request.POST.get('citizenship_number', user.citizenship_number)
                user.contact_number = request.POST.get('contact_number', user.contact_number)
                user.emergency_contact_number = request.POST.get('emergency_contact_number', user.emergency_contact_number)
                user.email = request.POST.get('email', user.email)
                user.address = request.POST.get('address', user.address)
                
                user.gender = request.POST.get('gender', user.gender)
                user.educational_qualifications = request.POST.get('educational_qualifications', user.educational_qualifications)
                user.languages_known = request.POST.get('languages_known', user.languages_known)
                user.proficiency_level = request.POST.get('proficiency_level', user.proficiency_level)
                
                # Update permissions
                user.global_telemetry = request.POST.get('global_telemetry') == 'true'
                user.ledger_override = request.POST.get('ledger_override') == 'true'
                user.api_key_gen = request.POST.get('api_key_gen') == 'true'
                user.audit_log_purge = request.POST.get('audit_log_purge') == 'true'
                user.hr_matrix_view = request.POST.get('hr_matrix_view') == 'true'
                user.fleet_command = request.POST.get('fleet_command') == 'true'
                
                # Page Access Permissions - Restricted to Admin
                active_role_upper = request.session.get('active_role', '').upper().strip()
                if active_role_upper in ['ADMIN', 'SYSTEM ADMIN', 'CEO', 'COO', 'CITO', 'HR AND OPERATION HEAD', 'HR AND OPERATION', 'HR', 'OPERATION HEAD', 'CHAIRMAN', 'OPERATION']:
                    if 'can_access_dashboard' in request.POST:
                        user.can_access_dashboard = request.POST.get('can_access_dashboard') == 'true'
                    if 'can_access_notice_board' in request.POST:
                        user.can_access_notice_board = request.POST.get('can_access_notice_board') == 'true'
                    if 'can_access_time_attendance' in request.POST:
                        user.can_access_time_attendance = request.POST.get('can_access_time_attendance') == 'true'
                    if 'can_access_leave' in request.POST:
                        user.can_access_leave = request.POST.get('can_access_leave') == 'true'
                    if 'can_access_profiles' in request.POST:
                        user.can_access_profiles = request.POST.get('can_access_profiles') == 'true'
                    if 'can_access_lead_pipeline' in request.POST:
                        user.can_access_lead_pipeline = request.POST.get('can_access_lead_pipeline') == 'true'
                    if 'can_access_inventory' in request.POST:
                        user.can_access_inventory = request.POST.get('can_access_inventory') == 'true'
                    if 'can_access_accounts_receivable' in request.POST:
                        user.can_access_accounts_receivable = request.POST.get('can_access_accounts_receivable') == 'true'
                    if 'can_access_pos' in request.POST:
                        user.can_access_pos = request.POST.get('can_access_pos') == 'true'
                    if 'can_access_procurement' in request.POST:
                        user.can_access_procurement = request.POST.get('can_access_procurement') == 'true'
                    if 'can_access_system_log' in request.POST:
                        user.can_access_system_log = request.POST.get('can_access_system_log') == 'true'
                    if 'can_access_task_board' in request.POST:
                        user.can_access_task_board = request.POST.get('can_access_task_board') == 'true'
                    if 'can_access_staff_payroll' in request.POST:
                        user.can_access_staff_payroll = request.POST.get('can_access_staff_payroll') == 'true'
                    if 'can_access_account_expenses' in request.POST:
                        user.can_access_account_expenses = request.POST.get('can_access_account_expenses') == 'true'
                
                user.save()
                
                if user.gender == 'Female':
                    from .models import CalendarEvent, LeaveRequest
                    female_holidays = CalendarEvent.objects.filter(is_female_holiday=True)
                    for fh in female_holidays:
                        LeaveRequest.objects.get_or_create(
                            employee_name=user.full_name,
                            leave_type='Casual',
                            start_date=fh.date,
                            end_date=fh.date,
                            defaults={
                                'reason': f'Female Holiday: {fh.title}',
                                'status': 'Approved'
                            }
                        )
                
                if request.session.get('logged_in_uid') == str(user.uid) or request.session.get('logged_in_name') == old_name:
                    request.session['logged_in_name'] = user.full_name
                    request.session['active_role'] = user.position
                    
                if old_name != user.full_name:
                    AttendanceRecord.objects.filter(employee_name=old_name).update(employee_name=user.full_name)
                    LeaveRequest.objects.filter(employee_name=old_name).update(employee_name=user.full_name)
                    
            if user:
                return redirect(f'/users/create/?uid={user.uid}')
            return redirect(f'/users/create/?uid={uid}')

        elif action == 'delete':
            active_role = request.session.get('active_role', 'Operations Lead').upper()
            if active_role in ['COO', 'CEO', 'HR AND OPERATION HEAD', 'CITO', 'ADMIN', 'SYSTEM ADMIN', 'HR', 'OPERATION HEAD']:
                SystemUserProfile.objects.filter(uid=uid).delete()
                return redirect('/users/create/')
            else:
                return redirect(f'/users/create/?uid={uid}&error=unauthorized')

        elif action == 'quick_provision':
            SystemUserProfile.objects.all().delete()
            return redirect('/users/create/')
            
        elif action == 'upload_document':
            document_title = request.POST.get('document_title')
            document_file = request.FILES.get('document_file')
            active_role_upper = request.session.get('active_role', '').upper()
            
            can_upload = active_role_upper in ['CEO', 'COO', 'HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD'] or uid == request.session.get('logged_in_uid')
            if can_upload and document_title and document_file:
                user = SystemUserProfile.objects.get(uid=uid)
                from .models import StaffDocument
                StaffDocument.objects.create(
                    user=user,
                    title=document_title,
                    file=document_file
                )
            return redirect(f'/users/create/?uid={uid}')

        return redirect('/users/create/')

from .models import Notice, LeaveRequest

class NoticeBoardView(TemplateView):
    template_name = "notice_board.html"

    def get(self, request, *args, **kwargs):
        notices = Notice.objects.all().order_by('-created_at')
        active_role = request.session.get('active_role', '')
        # Allow these roles to create notices
        can_create = active_role in ['CEO', 'COO', 'HR', 'Operation Head', 'CITO']
        can_delete = active_role in ['CEO', 'COO']
        return render(request, self.template_name, {
            'notices': notices,
            'can_create': can_create,
            'can_delete': can_delete,
            'active_role': active_role
        })

    def post(self, request, *args, **kwargs):
        active_role = request.session.get('active_role', '')
        
        # Handle Delete Action
        delete_notice_id = request.POST.get('delete_notice_id')
        if delete_notice_id and active_role in ['CEO', 'COO']:
            Notice.objects.filter(id=delete_notice_id).delete()
            return redirect('notice_board')

        # Handle Create/Edit Action
        if active_role in ['CEO', 'COO', 'HR', 'Operation Head', 'CITO']:
            action = request.POST.get('action', 'create')
            title = request.POST.get('title')
            content = request.POST.get('content')
            priority = request.POST.get('priority', 'Standard')
            department = request.POST.get('department', 'HQ')
            author_name = request.session.get('active_role', 'System Admin')
            
            if action == 'edit':
                edit_notice_id = request.POST.get('edit_notice_id')
                if edit_notice_id and title and content:
                    notice = Notice.objects.filter(id=edit_notice_id).first()
                    # Only allow edit if sender matches or user can delete (CEO/COO)
                    if notice and (notice.author_name == author_name or active_role in ['CEO', 'COO']):
                        notice.title = title
                        notice.content = content
                        notice.priority = priority
                        notice.department = department
                        notice.save()
            else:
                if title and content:
                    Notice.objects.create(
                        title=title,
                        content=content,
                        priority=priority,
                        department=department,
                        author_name=author_name
                    )
        return redirect('notice_board')


class LeaveListView(TemplateView):
    template_name = "leave.html"

    def get(self, request, *args, **kwargs):
        # Pre-populate default leave requests if none exist in database
        if not LeaveRequest.objects.exists():
            LeaveRequest.objects.create(
                employee_name="Hari Prasad",
                leave_type="Casual",
                start_date=date(2026, 6, 10),
                end_date=date(2026, 6, 12),
                reason="Family gathering in Pokhara",
                status="Approved"
            )
            LeaveRequest.objects.create(
                employee_name="John Doe",
                leave_type="Sick",
                start_date=date(2026, 6, 15),
                end_date=date(2026, 6, 15),
                reason="Dental checkup",
                status="Approved"
            )
            LeaveRequest.objects.create(
                employee_name="Sita Thapa",
                leave_type="Annual",
                start_date=date(2026, 6, 20),
                end_date=date(2026, 6, 25),
                reason="Visiting relatives abroad",
                status="Pending"
            )

        requests_list = LeaveRequest.objects.all().order_by('-created_at')

        from django.core.paginator import Paginator
        paginator = Paginator(requests_list, 25)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        # Allowances from DB
        from .models import LeaveSettings
        settings, created = LeaveSettings.objects.get_or_create(id=1)
        allowances = {
            'Casual': settings.casual_leave_days,
            'Sick': settings.sick_leave_days,
            'Annual': settings.annual_leave_days
        }
        
        # Calculate used days for approved leaves for John Doe (current operator)
        used = {k: 0 for k in allowances}
        for req in LeaveRequest.objects.filter(employee_name="John Doe", status="Approved"):
            duration = (req.end_date - req.start_date).days + 1
            if req.leave_type in used:
                used[req.leave_type] += duration

        balances = []
        for ltype, limit in allowances.items():
            used_days = used[ltype]
            remaining = max(0, limit - used_days)
            balances.append({
                'type': ltype,
                'name': dict(LeaveRequest.LEAVE_TYPES).get(ltype),
                'limit': limit,
                'used': used_days,
                'remaining': remaining,
                'percent': int((remaining / limit) * 100) if limit > 0 else 0
            })

        active_role = request.session.get('active_role', '')
        can_approve = active_role.upper() in ['COO', 'HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD', 'CEO', 'ADMIN', 'SYSTEM ADMIN', 'CITO']

        from .models import MissedAttendanceRequest
        missed_requests = MissedAttendanceRequest.objects.all().order_by('-created_at')

        context = {
            'leave_requests': page_obj,
            'balances': balances,
            'leave_types': LeaveRequest.LEAVE_TYPES,
            'can_approve': can_approve,
            'leave_settings': settings,
            'missed_requests': missed_requests
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        request_id = request.POST.get('request_id')

        if action == 'update_settings':
            active_role = request.session.get('active_role', '')
            user_name = request.session.get('logged_in_name', 'Unknown User')
            if active_role.upper() in ['COO', 'HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD', 'CEO', 'ADMIN', 'SYSTEM ADMIN', 'CITO']:
                from .models import LeaveSettings, SystemLog
                settings, _ = LeaveSettings.objects.get_or_create(id=1)
                try:
                    settings.casual_leave_days = int(request.POST.get('casual_leave_days', settings.casual_leave_days))
                    settings.sick_leave_days = int(request.POST.get('sick_leave_days', settings.sick_leave_days))
                    settings.annual_leave_days = int(request.POST.get('annual_leave_days', settings.annual_leave_days))
                    settings.maternity_leave_days = int(request.POST.get('maternity_leave_days', settings.maternity_leave_days))
                    settings.paternity_leave_days = int(request.POST.get('paternity_leave_days', settings.paternity_leave_days))
                    settings.save()
                    messages.success(request, "Global Leave Allowances updated successfully.")
                except ValueError:
                    messages.error(request, "Invalid input for leave allowances.")
            else:
                messages.error(request, "Access Denied: You do not have permission to edit leave settings.")
            return redirect('leave_list')

        if action in ['approve_missed', 'reject_missed'] and request_id:
            active_role = request.session.get('active_role', '').upper()
            if active_role in ['COO', 'HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD', 'CEO', 'ADMIN', 'SYSTEM ADMIN', 'CITO']:
                from .models import MissedAttendanceRequest, AttendanceRecord
                try:
                    ma_req = MissedAttendanceRequest.objects.get(id=request_id)
                    if action == 'reject_missed':
                        ma_req.status = 'Rejected'
                        ma_req.save()
                        messages.success(request, "Missed attendance request rejected.")
                    else:
                        if active_role == 'HR':
                            ma_req.hr_approved = True
                        elif active_role in ['OPERATION HEAD', 'HR AND OPERATION HEAD']:
                            ma_req.operation_approved = True
                        elif active_role == 'COO':
                            ma_req.coo_approved = True
                        elif active_role in ['CEO', 'ADMIN', 'SYSTEM ADMIN']:
                            ma_req.hr_approved = True
                            ma_req.operation_approved = True
                            ma_req.coo_approved = True
                            
                        if ma_req.is_fully_approved:
                            ma_req.status = 'Approved'
                            from datetime import timedelta
                            current_date = ma_req.start_date
                            while current_date <= ma_req.end_date:
                                AttendanceRecord.objects.update_or_create(
                                    employee_name=ma_req.employee.full_name,
                                    date=current_date,
                                    defaults={
                                        'check_in_time': ma_req.check_in_time,
                                        'check_out_time': ma_req.check_out_time,
                                        'status': 'Present'
                                    }
                                )
                                current_date += timedelta(days=1)
                        ma_req.save()
                        messages.success(request, "Missed attendance request approved.")
                except Exception as e:
                    messages.error(request, "Error processing missed attendance.")
            else:
                messages.error(request, "Access Denied.")
            return redirect('leave_list')

        if action in ['approve', 'reject'] and request_id:
            active_role = request.session.get('active_role', '')
            user_name = request.session.get('logged_in_name', 'Unknown User')
            if active_role.upper() in ['COO', 'HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD', 'CEO', 'ADMIN', 'SYSTEM ADMIN', 'CITO']:
                from .models import LeaveRequest, SystemLog
                try:
                    req = LeaveRequest.objects.get(id=request_id)
                    req.status = 'Approved' if action == 'approve' else 'Rejected'
                    req.save()
                    messages.success(request, f"Leave {req.status.lower()} for {req.employee_name}. Notification & Email sent to COO, HR, Operation Head, and Applicant.")
                except LeaveRequest.DoesNotExist:
                    pass
            else:
                messages.error(request, "Access Denied: Only COO, HR, and Operation Head can approve leaves.")
        else:
            from .models import LeaveRequest
            employee_name = request.session.get('logged_in_name') or request.POST.get('employee_name', 'John Doe')
            leave_type = request.POST.get('leave_type')
            start_date_str = request.POST.get('start_date')
            end_date_str = request.POST.get('end_date')
            reason = request.POST.get('reason', '')

            if leave_type and start_date_str and end_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                    LeaveRequest.objects.create(
                        employee_name=employee_name,
                        leave_type=leave_type,
                        start_date=start_date,
                        end_date=end_date,
                        reason=reason,
                        status='Pending'
                    )
                    messages.success(request, f"Leave requested! Notification & Email sent to COO, HR, and Operation Head.")
                except Exception as e:
                    messages.error(request, f"Failed to create leave request: {str(e)} (Dates must be YYYY-MM-DD)")
            else:
                messages.error(request, "Please fill out all required fields.")

        return redirect('leave_list')


from .models import LeadCRM, OrgNode

class LeadPipelineView(TemplateView):
    template_name = "leads.html"

    def get(self, request, *args, **kwargs):
        # Ensure default nodes exist
        if not OrgNode.objects.exists():
            hq = OrgNode.objects.create(name="HQ-NEPAL", node_type="HQ")
            dist = OrgNode.objects.create(name="DISTRIBUTOR-BAGMATI", node_type="CAT_A", parent=hq)
            OrgNode.objects.create(name="DEALER-KATHMANDU", node_type="CAT_B", parent=dist)

        default_node = OrgNode.objects.first()

        # Ensure default leads exist
        if not LeadCRM.objects.exists():
            LeadCRM.objects.create(lead_name="Ncell Axiata", node=default_node, pipeline_state="Proposal", phone="+977-9801200000", address="Lalitpur, Nepal", deal_value=250000.0, product_inquiry="Cisco ISR 4321 Router")
            LeadCRM.objects.create(lead_name="Nepal Telecom", node=default_node, pipeline_state="Identified", phone="+977-1-4210000", address="Bhadrakali, Kathmandu", deal_value=450000.0, product_inquiry="Hikvision PTZ Cameras")
            LeadCRM.objects.create(lead_name="WorldLink Communications", node=default_node, pipeline_state="Converted_To_Project", phone="+977-1-5970050", address="Jawalakhel, Lalitpur", deal_value=600000.0, product_inquiry="Ubiquiti UniFi APs")
            LeadCRM.objects.create(lead_name="Subisu Cablenet", node=default_node, pipeline_state="Dead", phone="+977-1-4429616", address="Baluwatar, Kathmandu", deal_value=120000.0, product_inquiry="Dahua DVR 16-ch")
            LeadCRM.objects.create(lead_name="DishHome Cable", node=default_node, pipeline_state="Proposal", phone="+977-1-5970000", address="Chabahil, Kathmandu", deal_value=180000.0, product_inquiry="Fiber Optic Spools 10km")
            LeadCRM.objects.create(lead_name="Vianet Communications", node=default_node, pipeline_state="Identified", phone="+977-1-5970555", address="Jawalakhel, Lalitpur", deal_value=320000.0, product_inquiry="MikroTik Cloud Core Router")

        leads = LeadCRM.objects.all().order_by('-created_at')
        nodes = OrgNode.objects.all()

        # Calculate metrics dynamically
        total_leads = leads.count()
        converted_leads = leads.filter(pipeline_state="Converted_To_Project").count()
        dead_leads = leads.filter(pipeline_state="Dead").count()

        win_rate = (converted_leads / total_leads * 100) if total_leads > 0 else 0.0
        dead_ratio = (dead_leads / total_leads * 100) if total_leads > 0 else 0.0

        from django.db.models import Sum
        total_value = leads.aggregate(total=Sum('deal_value'))['total'] or 0.0

        stages = {
            'Identified': [],
            'Proposal': [],
            'Converted_To_Project': [],
            'Dead': []
        }
        for lead in leads:
            if lead.pipeline_state in stages:
                stages[lead.pipeline_state].append(lead)

        context = {
            'stages': stages,
            'nodes': nodes,
            'total_leads': total_leads,
            'win_rate': round(win_rate, 1),
            'dead_ratio': round(dead_ratio, 1),
            'total_value': total_value,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        lead_id = request.POST.get('lead_id')

        if action == 'create':
            lead_name = request.POST.get('lead_name')
            node_id = request.POST.get('node_id')
            pipeline_state = request.POST.get('pipeline_state', 'Identified')
            phone = request.POST.get('phone', '')
            address = request.POST.get('address', '')
            deal_value_str = request.POST.get('deal_value', '0')
            product_inquiry = request.POST.get('product_inquiry', '')
            remarks = request.POST.get('remarks', '')
            try:
                deal_value = float(deal_value_str)
            except ValueError:
                deal_value = 0.0

            if lead_name and node_id:
                try:
                    node = OrgNode.objects.get(id=node_id)
                    LeadCRM.objects.create(
                        lead_name=lead_name,
                        node=node,
                        pipeline_state=pipeline_state,
                        phone=phone,
                        address=address,
                        deal_value=deal_value,
                        product_inquiry=product_inquiry,
                        remarks=remarks
                    )
                except OrgNode.DoesNotExist:
                    pass

        elif action == 'update_stage':
            new_stage = request.POST.get('pipeline_state')
            if lead_id and new_stage:
                try:
                    lead = LeadCRM.objects.get(id=lead_id)
                    lead.pipeline_state = new_stage
                    lead.save()
                except LeadCRM.DoesNotExist:
                    pass

        elif action == 'delete':
            if lead_id:
                LeadCRM.objects.filter(id=lead_id).delete()

        return redirect('lead_pipeline')

import csv
from django.http import HttpResponse
from django.views import View

class LeadReportExcelView(View):
    def get(self, request, *args, **kwargs):
        from .models import LeadCRM
        import openpyxl
        from django.http import HttpResponse

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Leads Report'

        columns = ['Lead ID', 'Lead Name', 'Node', 'Pipeline State', 'Phone', 'Address', 'Deal Value', 'Product Inquiry', 'Remarks', 'Created At']
        worksheet.append(columns)
        
        leads = LeadCRM.objects.all().order_by('-created_at')
        for lead in leads:
            worksheet.append([
                lead.id, 
                lead.lead_name, 
                lead.node.name if lead.node else 'N/A', 
                lead.pipeline_state, 
                lead.phone, 
                lead.address, 
                float(lead.deal_value), 
                lead.product_inquiry, 
                lead.remarks,
                lead.created_at.strftime("%Y-%m-%d")
            ])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="leads_report.xlsx"'
        workbook.save(response)
        return response


from django.views import View
from .models import ProcurementRequest
from .services import ProcurementStateMachine

class ProcurementView(View):
    template_name = "procurement.html"

    def get(self, request, *args, **kwargs):
        # Ensure default procurement requests exist for testing
        if not ProcurementRequest.objects.exists():
            ProcurementRequest.objects.create(title="Purchase 15x CCTV Cameras (Hikvision)", track="Local", hr_approved=True, coo_approved=False)
            ProcurementRequest.objects.create(title="Import Server Rack Infrastructure", track="International", hr_approved=True, cto_approved=True, coo_approved=False)
            ProcurementRequest.objects.create(title="Office Chairs & Desks Refurbishment", track="Local", hr_approved=False, coo_approved=False)
            ProcurementRequest.objects.create(title="Nightvision AI Drone Prototype Import", track="International", hr_approved=True, cto_approved=True, coo_approved=True, ceo_approved=True, is_terminal=True)

        requests = ProcurementRequest.objects.all().order_by('-id')

        # Dynamically enrich each request in python with the "next_signer" role
        for req in requests:
            if req.is_terminal:
                req.next_signer = None
            elif req.track == 'International':
                if not req.hr_approved:
                    req.next_signer = 'HR'
                elif not req.cto_approved:
                    req.next_signer = 'CTO'
                elif not req.coo_approved:
                    req.next_signer = 'COO'
                elif not req.ceo_approved:
                    req.next_signer = 'CEO'
                else:
                    req.next_signer = None
            elif req.track == 'Local':
                if not req.hr_approved:
                    req.next_signer = 'HR'
                elif not req.coo_approved:
                    req.next_signer = 'COO'
                else:
                    req.next_signer = None

        # Statistics
        total_count = requests.count()
        completed_count = requests.filter(is_terminal=True).count()
        pending_local = requests.filter(track='Local', is_terminal=False).count()
        pending_intl = requests.filter(track='International', is_terminal=False).count()

        context = {
            'requests': requests,
            'total_count': total_count,
            'completed_count': completed_count,
            'pending_local': pending_local,
            'pending_intl': pending_intl,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        req_id = request.POST.get('procurement_id')
        user_uid = request.session.get('logged_in_uid')

        if action == 'create':
            title = request.POST.get('title')
            track = request.POST.get('track', 'Local')
            supporting_document = request.FILES.get('supporting_document')
            landing_cost_management = request.POST.get('landing_cost_management', '0.00')
            goods_receive_notes = request.POST.get('goods_receive_notes')
            if title:
                lcm = 0.00
                if landing_cost_management:
                    try:
                        lcm = float(landing_cost_management)
                    except ValueError:
                        pass
                ProcurementRequest.objects.create(
                    title=title, track=track, supporting_document=supporting_document,
                    landing_cost_management=lcm, goods_receive_notes=goods_receive_notes,
                    submitted_by_uid=user_uid
                )
        
        elif action == 'approve':
            role = request.POST.get('role')
            if req_id and role:
                user_role = request.session.get('active_role', '')
                success, msg = ProcurementStateMachine.approve_step(req_id, role, user_uid, user_role)
                if not success:
                    from django.contrib import messages
                    messages.error(request, msg)
                else:
                    from django.contrib import messages
                    messages.success(request, msg)
        
        elif action == 'delete':
            if req_id:
                ProcurementRequest.objects.filter(id=req_id).delete()

        return redirect('procurement_list')


from .models import AccountsReceivable, OrgNode
from datetime import date, datetime, timedelta

class AccountsReceivableView(View):
    template_name = "accounts_receivable.html"

    def get(self, request, *args, **kwargs):
        # Pre-populate default accounts receivable if none exist in database
        if not AccountsReceivable.objects.exists():
            # Get default nodes
            hq = OrgNode.objects.filter(name="HQ-NEPAL").first()
            if not hq:
                hq = OrgNode.objects.create(name="HQ-NEPAL", node_type="HQ")
            dist = OrgNode.objects.filter(name="DISTRIBUTOR-BAGMATI").first() or OrgNode.objects.create(name="DISTRIBUTOR-BAGMATI", node_type="CAT_A", parent=hq)
            dealer = OrgNode.objects.filter(name="DEALER-KATHMANDU").first() or OrgNode.objects.create(name="DEALER-KATHMANDU", node_type="CAT_B", parent=dist)
            
            AccountsReceivable.objects.create(
                invoice_number="INV-2026-001",
                node=hq,
                amount_due=24500.00,
                amount_paid=0.0,
                state="Unpaid",
                due_date=date.today() + timedelta(days=15)
            )
            AccountsReceivable.objects.create(
                invoice_number="INV-2026-002",
                node=dist,
                amount_due=112000.00,
                amount_paid=112000.00,
                state="Settled",
                due_date=date.today() - timedelta(days=5)
            )
            AccountsReceivable.objects.create(
                invoice_number="INV-2026-003",
                node=dealer,
                amount_due=4120.50,
                amount_paid=2000.00,
                state="Partially_Paid",
                due_date=date.today() + timedelta(days=30)
            )
            AccountsReceivable.objects.create(
                invoice_number="INV-2026-004",
                node=dist,
                amount_due=67900.00,
                amount_paid=0.0,
                state="Disputed",
                due_date=date.today() + timedelta(days=2)
            )

        invoices = AccountsReceivable.objects.all().order_by('-created_at')
        nodes = OrgNode.objects.all()

        # Calculate metrics dynamically
        total_invoiced = 0.0
        total_settled = 0.0
        total_receivables = 0.0
        critical_past_due = 0.0

        today = date.today()

        for inv in invoices:
            due = float(inv.amount_due)
            paid = float(inv.amount_paid)
            outstanding = max(0.0, due - paid)

            total_invoiced += due
            total_settled += paid
            total_receivables += outstanding
                
            # Critical Past Due is defined as Unpaid/Disputed/Partially Paid invoices past due date
            if inv.state in ['Unpaid', 'Disputed', 'Partially_Paid'] and inv.due_date and inv.due_date < today:
                critical_past_due += outstanding

        active_role = request.session.get('active_role', '')
        is_admin = active_role in ['CEO', 'COO', 'HR', 'Operation Head', 'CITO']
        
        context = {
            'invoices': invoices,
            'nodes': nodes,
            'total_invoiced': total_invoiced,
            'total_settled': total_settled,
            'total_receivables': total_receivables,
            'critical_past_due': critical_past_due,
            'is_admin': is_admin,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        active_role = request.session.get('active_role', '')
        is_admin = active_role in ['CEO', 'COO', 'HR', 'Operation Head', 'CITO']
        
        action = request.POST.get('action')
        invoice_id = request.POST.get('invoice_id')
        
        if action == 'delete' and invoice_id:
            system_user = getattr(request, 'system_user', None)
            is_ceo = (active_role == 'CEO') or (system_user and system_user.position == 'CEO')
            
            if is_ceo:
                AccountsReceivable.objects.filter(id=invoice_id).delete()
                from django.contrib import messages
                messages.success(request, 'Invoice deleted successfully.')
            else:
                from django.contrib import messages
                messages.error(request, 'Access Denied: Only the CEO can delete generated invoices.')
            return redirect('accounts_receivable')

        if action == 'create':
            invoice_number = request.POST.get('invoice_number')
            node_id = request.POST.get('node_id')
            amount_due_str = request.POST.get('amount_due', '0')
            amount_paid_str = request.POST.get('amount_paid', '0')
            state = request.POST.get('state', 'Unpaid')
            due_date_str = request.POST.get('due_date')

            try:
                amount_due = float(amount_due_str)
            except ValueError:
                amount_due = 0.0

            try:
                amount_paid = float(amount_paid_str)
            except ValueError:
                amount_paid = 0.0

            if state == 'Settled':
                amount_paid = amount_due
            elif state in ['Unpaid', 'Disputed'] and not amount_paid_str:
                amount_paid = 0.0

            due_date_val = None
            if due_date_str:
                try:
                    import nepali_datetime
                    year_part, month_part, day_part = map(int, due_date_str.split('-'))
                    np_date = nepali_datetime.date(year_part, month_part, day_part)
                    due_date_val = np_date.to_datetime_date()
                except Exception:
                    try:
                        due_date_val = datetime.strptime(due_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        pass

            if invoice_number and node_id:
                try:
                    node = OrgNode.objects.get(id=node_id)
                    AccountsReceivable.objects.create(
                        invoice_number=invoice_number,
                        node=node,
                        amount_due=amount_due,
                        amount_paid=amount_paid,
                        state=state,
                        due_date=due_date_val
                    )
                except OrgNode.DoesNotExist:
                    pass

        elif action == 'update':
            if invoice_id:
                try:
                    inv = AccountsReceivable.objects.get(id=invoice_id)
                    invoice_number = request.POST.get('invoice_number')
                    node_id = request.POST.get('node_id')
                    amount_due_str = request.POST.get('amount_due')
                    amount_paid_str = request.POST.get('amount_paid')
                    state = request.POST.get('state')
                    due_date_str = request.POST.get('due_date')

                    if invoice_number:
                        inv.invoice_number = invoice_number
                    if node_id:
                        try:
                            inv.node = OrgNode.objects.get(id=node_id)
                        except OrgNode.DoesNotExist:
                            pass
                    if amount_due_str is not None:
                        try:
                            inv.amount_due = float(amount_due_str)
                        except ValueError:
                            pass
                    
                    if state:
                        inv.state = state

                    if amount_paid_str is not None:
                        try:
                            inv.amount_paid = float(amount_paid_str)
                        except ValueError:
                            pass

                    # Auto correction logic based on state changes
                    if inv.state == 'Settled':
                        inv.amount_paid = inv.amount_due
                    elif inv.state in ['Unpaid', 'Disputed'] and amount_paid_str is None:
                        inv.amount_paid = 0.0

                    if due_date_str:
                        try:
                            import nepali_datetime
                            year_part, month_part, day_part = map(int, due_date_str.split('-'))
                            np_date = nepali_datetime.date(year_part, month_part, day_part)
                            inv.due_date = np_date.to_datetime_date()
                        except Exception:
                            try:
                                inv.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
                            except ValueError:
                                pass
                    inv.save()
                except AccountsReceivable.DoesNotExist:
                    pass

        elif action == 'delete':
            pass # Deletion is now fully handled at the beginning of the post method

        return redirect('accounts_receivable')


class SwitchRoleView(View):
    def post(self, request, *args, **kwargs):
        active_role = request.POST.get('active_role', 'Operations Lead')
        request.session['active_role'] = active_role
        return redirect(request.META.get('HTTP_REFERER', '/'))

from django.db.models import Q
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache

@method_decorator(never_cache, name='dispatch')
class LoginView(View):
    def get(self, request, *args, **kwargs):
        if request.session.get('logged_in_uid'):
            return redirect('dashboard')
        return render(request, 'login.html')

    def post(self, request, *args, **kwargs):
        login_id = request.POST.get('login_id', '').strip()
        password = request.POST.get('password', '')
        
        user = SystemUserProfile.objects.filter(Q(uid__iexact=login_id) | Q(full_name__iexact=login_id)).first()
        
        if user and user.password == password:
            request.session['logged_in_uid'] = user.uid
            request.session['logged_in_name'] = user.full_name
            request.session['active_role'] = user.position
            return redirect('dashboard')
        else:
            return render(request, 'login.html', {'error': 'Invalid Identity Token or Access Key.'})

@method_decorator(never_cache, name='dispatch')
class LogoutView(View):
    def get(self, request, *args, **kwargs):
        request.session.flush()
        return redirect('login')

import json
import uuid
from django.contrib import messages
from django.views.generic import View

class POSView(TemplateView):
    template_name = "pos.html"

    def get(self, request, *args, **kwargs):
        from .models import InventoryLedger
        inventory_items = []
        for item in InventoryLedger.objects.all():
            inventory_items.append({
                'stock_record_id': str(item.stock_record_id),
                'sku_id': item.sku_id,
                'product_name': item.product_name,
                'product_image': item.product_image.url if item.product_image else None,
                'quantity_on_hand': item.quantity_on_hand,
                'price': float(item.price)
            })
        
        return render(request, self.template_name, {'inventory': inventory_items})

class POSCheckoutView(View):
    def post(self, request, *args, **kwargs):
        from .models import InventoryLedger, AccountsReceivable, OrgNode
        cart_data_str = request.POST.get('cart_data', '{}')
        total_amount = request.POST.get('total_amount', '0.00')
        discount = request.POST.get('discount', '0')
        customer_name = request.POST.get('customer_name', 'WALK-IN CUSTOMER')
        customer_pan = request.POST.get('customer_pan', '')
        customer_phone = request.POST.get('customer_phone', '')
        customer_address = request.POST.get('customer_address', '')
        payment_method = request.POST.get('payment_method', 'Cash')
        
        try:
            cart = json.loads(cart_data_str)
            
            hq_node = OrgNode.objects.filter(node_type='HQ').first()
            if not hq_node:
                hq_node = OrgNode.objects.create(name="HQ-NEPAL-POS", node_type="HQ")
                
            invoice_num = f"POS-{str(uuid.uuid4())[:8].upper()}"
            purchased_items = []

            # 1. Deduct Inventory
            for record_id, item_data in cart.items():
                qty = item_data.get('qty', 0)
                inv_item = InventoryLedger.objects.get(stock_record_id=record_id)
                if inv_item.quantity_on_hand >= qty:
                    inv_item.quantity_on_hand -= qty
                    inv_item.save()
                    purchased_items.append({
                        'sku_id': inv_item.sku_id,
                        'product_name': inv_item.product_name,
                        'quantity': qty,
                        'price': float(inv_item.price),
                        'total': qty * float(inv_item.price)
                    })
            
            # 2. Record in Accounts Receivable
            paid_amount = total_amount if payment_method in ['Cash', 'Cheque'] else 0.0
            ar_state = 'Settled' if payment_method in ['Cash', 'Cheque'] else 'Unpaid'
            
            AccountsReceivable.objects.create(
                invoice_number=invoice_num,
                customer_name=customer_name,
                customer_pan=customer_pan,
                customer_phone=customer_phone,
                customer_address=customer_address,
                node=hq_node,
                amount_due=total_amount,
                amount_paid=paid_amount,
                state=ar_state,
                payment_method=payment_method
            )
            
            from django.utils import timezone
            # Save invoice to session for printable page
            request.session['last_invoice'] = {
                'invoice_id': invoice_num,
                'date': timezone.localtime().strftime("%Y-%m-%d %H:%M"),
                'customer_name': customer_name,
                'customer_pan': customer_pan,
                'customer_phone': customer_phone,
                'customer_address': customer_address,
                'payment_method': payment_method,
                'items': purchased_items,
                'subtotal': sum(i['total'] for i in purchased_items),
                'discount': float(discount),
                'total_amount': float(total_amount)
            }
            
            messages.success(request, f"Checkout complete! Invoice {invoice_num} generated.")
            return redirect('pos_invoice')
            
        except Exception as e:
            messages.error(request, f"Checkout failed: {str(e)}")
            return redirect('pos')

class POSInvoiceView(TemplateView):
    template_name = "pos_invoice.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['invoice'] = self.request.session.get('last_invoice', None)
        return context

class StaffPayrollReportView(TemplateView):
    template_name = "attendance_payroll_report.html"

    def get(self, request, *args, **kwargs):
        from django.core.exceptions import PermissionDenied
        from .models import EmployeeProfile, AttendanceRecord, SystemUserProfile
        import csv
        from django.http import HttpResponse
        
        active_role = request.session.get('active_role', '').upper()
        allowed_roles = ['CEO', 'COO', 'HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD', 'CITO', 'ADMIN', 'CHAIRMAN', 'OPERATION']
        
        if active_role not in allowed_roles:
            messages.error(request, "Access Denied: You do not have permission to view the Payroll Report.")
            return redirect('dashboard')
            
        # 1. Fetch Attendance Records for Display
        attendance_records = AttendanceRecord.objects.all().order_by('-date', '-created_at')
        
        # 2. Fetch all Staff/System Users
        users = SystemUserProfile.objects.all()
        
        # 3. Calculate Payroll per User
        payroll_data = []
        for u in users:
            # Let's try to find an EmployeeProfile for their gross salary, otherwise default to 50000
            emp = EmployeeProfile.objects.filter(first_name__icontains=u.full_name.split(' ')[0]).first()
            base_salary = float(emp.base_gross_salary) if emp else 50000.0
            
            # Day-based calculations for Excel format
            emp_records = AttendanceRecord.objects.filter(employee_name=u.full_name, status__in=['Present', 'Late'])
            
            att_days = len(set(r.date for r in emp_records))
            dearness = float(emp.dearness) if emp else 0.0
            allowance = float(emp.allowance) if emp else 0.0
            per_day = base_salary / 30.0 if base_salary else 0.0
            total_gross = (att_days * per_day) + dearness + allowance
            sst = total_gross * 0.01
            rem_tax = 0.0
            total_tds = sst + rem_tax
            net_payable = total_gross - total_tds
            
            payroll_data.append({
                'employee_name': u.full_name,
                'position': u.position,
                'att_days': att_days,
                'base_salary': round(base_salary, 2),
                'dearness': dearness,
                'allowance': allowance,
                'per_day': round(per_day, 2),
                'total_gross': round(total_gross, 2),
                'sst': round(sst, 2),
                'rem_tax': round(rem_tax, 2),
                'total_tds': round(total_tds, 2),
                'net_payable': round(net_payable, 2),
                'records': emp_records.order_by('date')
            })
            
        # Additional default stats for John Doe if the users table doesn't match him
        if not any(pd['employee_name'] == 'John Doe' for pd in payroll_data):
            john_records = AttendanceRecord.objects.filter(employee_name='John Doe', status='Present')
            if john_records.exists():
                john_days = len(set(r.date for r in john_records))
                j_base = 50000.0
                j_per_day = j_base / 30.0
                j_gross = john_days * j_per_day
                j_sst = j_gross * 0.01
                j_tds = j_sst
                j_net = j_gross - j_tds
                
                payroll_data.append({
                    'employee_name': 'John Doe',
                    'position': 'Staff',
                    'att_days': john_days,
                    'base_salary': j_base,
                    'dearness': 0.0,
                    'allowance': 0.0,
                    'per_day': round(j_per_day, 2),
                    'total_gross': round(j_gross, 2),
                    'sst': round(j_sst, 2),
                    'rem_tax': 0.0,
                    'total_tds': round(j_tds, 2),
                    'net_payable': round(j_net, 2),
                    'records': john_records.order_by('date')
                })

        export_user = request.GET.get('export_user')
        if export_user:
            employee_data = None
            for row in payroll_data:
                if row['employee_name'] == export_user:
                    employee_data = row
                    break
                    
            if employee_data:
                from .models import SystemLog
                SystemLog.objects.create(
                    user_name=request.session.get('logged_in_name', 'Unknown'),
                    action="Generated Individual PDF Payslip",
                    details=f"Generated printable PDF payslip for {export_user}."
                )
                from django.utils import timezone
                return render(request, 'payslip_print.html', {
                    'employee_data': employee_data,
                    'now': timezone.now()
                })

        from django.core.paginator import Paginator
        paginator = Paginator(attendance_records, 25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'attendance_records': page_obj,
            'payroll_data': payroll_data
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        from .models import EmployeeProfile, SystemUserProfile, OrgNode
        active_role = request.session.get('active_role', '').upper()
        allowed_roles = ['CEO', 'COO', 'HR AND OPERATION HEAD']
        
        if active_role not in allowed_roles:
            messages.error(request, "Access Denied: You do not have permission to edit payroll.")
            return redirect('attendance_payroll_report')
            
        action = request.POST.get('action')
        if action == 'update_salary':
            employee_name = request.POST.get('employee_name')
            new_salary = request.POST.get('new_salary')
            new_dearness = request.POST.get('dearness', 0.0)
            new_allowance = request.POST.get('allowance', 0.0)
            
            try:
                new_salary = float(new_salary)
                new_dearness = float(new_dearness)
                new_allowance = float(new_allowance)
                first_name = employee_name.split(' ')[0]
                emp = EmployeeProfile.objects.filter(first_name__icontains=first_name).first()
                if emp:
                    emp.base_gross_salary = new_salary
                    emp.dearness = new_dearness
                    emp.allowance = new_allowance
                    emp.save()
                else:
                    # Create one if doesn't exist to bind the salary
                    default_node = OrgNode.objects.first()
                    parts = employee_name.split(' ', 1)
                    last_name = parts[1] if len(parts) > 1 else ''
                    EmployeeProfile.objects.create(
                        first_name=first_name,
                        last_name=last_name,
                        node=default_node,
                        base_gross_salary=new_salary,
                        dearness=new_dearness,
                        allowance=new_allowance
                    )
                messages.success(request, f"Payroll for {employee_name} updated successfully.")
            except ValueError:
                messages.error(request, "Invalid amounts provided.")
                
        return redirect('attendance_payroll_report')

class SystemLogView(TemplateView):
    template_name = "system_log.html"

    def get(self, request, *args, **kwargs):
        active_role = request.session.get('active_role', '')
        if active_role.upper() not in ['COO', 'HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD', 'CEO', 'ADMIN', 'SYSTEM ADMIN', 'CITO', 'CHAIRMAN', 'OPERATION']:
            messages.error(request, "Access Denied: You do not have permission to view System Logs.")
            return redirect('dashboard')
            
        from .models import SystemLog
        from django.core.paginator import Paginator
        
        log_list = SystemLog.objects.all().order_by('-timestamp')
        paginator = Paginator(log_list, 25)
        page_number = request.GET.get('page')
        logs = paginator.get_page(page_number)
        
        return render(request, self.template_name, {'logs': logs})

class SystemLogExcelView(View):
    def get(self, request, *args, **kwargs):
        active_role = request.session.get('active_role', '')
        if active_role.upper() not in ['COO', 'HR', 'OPERATION HEAD', 'HR AND OPERATION HEAD', 'CEO', 'ADMIN', 'SYSTEM ADMIN', 'CITO', 'CHAIRMAN', 'OPERATION']:
            messages.error(request, "Access Denied: You do not have permission to download System Logs.")
            return redirect('dashboard')
            
        from .models import SystemLog
        import openpyxl
        from django.http import HttpResponse
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'System Logs'
        
        worksheet.append(['Timestamp', 'User', 'Action', 'Details'])
        
        logs = SystemLog.objects.all().order_by('-timestamp')
        for log in logs:
            worksheet.append([
                log.timestamp.strftime("%b %d, %Y %H:%M") if log.timestamp else 'N/A',
                log.user_name,
                log.action,
                log.details
            ])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="system_logs.xlsx"'
        workbook.save(response)
        return response

class ProjectTaskBoardView(TemplateView):
    template_name = "task_board.html"

    def get(self, request, *args, **kwargs):
        from .models import ProjectTask, SystemUserProfile
        
        active_role = request.session.get('active_role', '').upper()
        logged_in_name = request.session.get('logged_in_name', '')
        management_roles = ['CEO', 'COO', 'HR AND OPERATION HEAD', 'HR', 'OPERATION HEAD', 'CITO', 'CSO', 'ADMIN', 'SYSTEM ADMIN']
        is_management = active_role in management_roles
        
        tasks = ProjectTask.objects.all().order_by('-created_at')
        view_user_name = request.GET.get('view_user_name', '')
        
        if is_management:
            if view_user_name:
                tasks = tasks.filter(assigned_to__full_name=view_user_name)
        else:
            # Normal users can only see tasks assigned to them
            tasks = tasks.filter(assigned_to__full_name=logged_in_name)
            
        total_tasks = tasks.count()
        
        # Group tasks by status
        kanban_data = {
            'TODO': tasks.filter(status='TODO'),
            'IN_PROGRESS': tasks.filter(status='IN_PROGRESS'),
            'REVIEW': tasks.filter(status='REVIEW'),
            'DONE': tasks.filter(status='DONE'),
        }
        
        metrics = {
            'todo_pct': 0,
            'in_progress_pct': 0,
            'review_pct': 0,
            'done_pct': 0,
            'total': total_tasks
        }
        
        if total_tasks > 0:
            metrics['todo_pct'] = round((kanban_data['TODO'].count() / total_tasks) * 100)
            metrics['in_progress_pct'] = round((kanban_data['IN_PROGRESS'].count() / total_tasks) * 100)
            metrics['review_pct'] = round((kanban_data['REVIEW'].count() / total_tasks) * 100)
            metrics['done_pct'] = round((kanban_data['DONE'].count() / total_tasks) * 100)
        
        # Get all users for the assignment dropdown
        users = SystemUserProfile.objects.all()
        
        can_delete = active_role in ['CEO', 'COO', 'HR AND OPERATION HEAD', 'HR', 'OPERATION HEAD', 'CITO']
        context = {
            'kanban_data': kanban_data,
            'metrics': metrics,
            'users': users,
            'today': date.today().strftime('%Y-%m-%d'),
            'is_management': is_management,
            'view_user_name': view_user_name,
            'can_delete': can_delete
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        from .models import ProjectTask, SystemUserProfile, SystemNotification
        
        action = request.POST.get('action')
        logged_in_name = request.session.get('logged_in_name')
        
        if action == 'create_task':
            title = request.POST.get('title')
            description = request.POST.get('description', '')
            priority = request.POST.get('priority', 'MEDIUM')
            assigned_to_id = request.POST.get('assigned_to')
            due_date = request.POST.get('due_date')
            
            assigned_to = None
            if assigned_to_id:
                assigned_to = SystemUserProfile.objects.filter(id=assigned_to_id).first()
                
            assigned_by = None
            if logged_in_name:
                assigned_by = SystemUserProfile.objects.filter(full_name=logged_in_name).first()
                
            if due_date:
                try:
                    import nepali_datetime
                    parts = due_date.split('-')
                    if len(parts) == 3:
                        np_date = nepali_datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
                        due_date = np_date.to_datetime_date()
                except Exception:
                    pass
                    
            task = ProjectTask.objects.create(
                title=title,
                description=description,
                priority=priority,
                assigned_to=assigned_to,
                assigned_by=assigned_by,
                due_date=due_date if due_date else None
            )
            
            # Send Notification to Assignee
            if assigned_to and assigned_by and assigned_to.id != assigned_by.id:
                SystemNotification.objects.create(
                    recipient=assigned_to,
                    message=f"{assigned_by.full_name} assigned you a new task: '{title}'",
                    link="/task-board/"
                )
                
            messages.success(request, 'Task created successfully.')
            
        elif action == 'delete_task':
            task_id = request.POST.get('task_id')
            active_role = request.session.get('active_role', '').upper()
            allowed_roles = ['CEO', 'COO', 'HR AND OPERATION HEAD', 'HR', 'OPERATION HEAD', 'CITO']
            if active_role in allowed_roles:
                task = ProjectTask.objects.filter(id=task_id).first()
                if task:
                    task.delete()
                    from django.contrib import messages
                    messages.success(request, 'Task deleted successfully.')
            else:
                from django.contrib import messages
                messages.error(request, 'Access Denied: You do not have permission to delete tasks.')

        elif action == 'update_status':
            task_id = request.POST.get('task_id')
            new_status = request.POST.get('new_status')
            
            if task_id and new_status:
                task = ProjectTask.objects.filter(id=task_id).first()
                if task:
                    task.status = new_status
                    task.save()
                    
                    # Send Notification to Creator
                    if task.assigned_by and task.assigned_by.full_name != logged_in_name:
                        # Get friendly status name
                        status_display = dict(ProjectTask.STATUS_CHOICES).get(new_status, new_status)
                        SystemNotification.objects.create(
                            recipient=task.assigned_by,
                            message=f"{logged_in_name} moved '{task.title}' to {status_display}",
                            link="/task-board/"
                        )
                        
                    from django.http import JsonResponse
                    return JsonResponse({'success': True})
                    
        return redirect('task_board')

from django.http import JsonResponse

class NoticePollView(View):
    def get(self, request):
        last_check_str = request.GET.get('last_check')
        if not last_check_str:
            return JsonResponse({'notices': []})
            
        try:
            # Handle standard ISO format from JS
            last_check = datetime.fromisoformat(last_check_str.replace('Z', '+00:00'))
        except Exception:
            return JsonResponse({'notices': []})
            
        # Get notices created AFTER the last_check time
        new_notices = Notice.objects.filter(created_at__gt=last_check).order_by('created_at')
        
        data = []
        for n in new_notices:
            data.append({
                'title': n.title,
                'author': n.author_name,
                'priority': n.priority,
                'content': n.content[:100] + '...' if len(n.content) > 100 else n.content
            })
            
        return JsonResponse({'notices': data})

from django.db.models import Sum
from .models import ExpenseRecord

def account_expenses(request):
    if 'logged_in_uid' not in request.session:
        return redirect('login_view')
    
    uid = request.session['logged_in_uid']
    system_user = SystemUserProfile.objects.filter(uid=uid).first()
    
    if system_user and not system_user.can_access_account_expenses:
        return redirect('dashboard')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_expense':
            title = request.POST.get('title')
            amount = request.POST.get('amount')
            category = request.POST.get('category')
            date_val = request.POST.get('date')
            import nepali_datetime
            if not date_val:
                date_val = str(nepali_datetime.date.today())
            description = request.POST.get('description', '')
            
            ExpenseRecord.objects.create(
                title=title,
                amount=amount,
                category=category,
                date=date_val,
                description=description,
                logged_by=system_user
            )
            messages.success(request, f"Successfully logged expense: Rs. {amount} for {title}")
            return redirect('account_expenses')
            
        elif action == 'delete_expense':
            if request.session.get('active_role') == 'CEO':
                expense_id = request.POST.get('expense_id')
                if expense_id:
                    expense = ExpenseRecord.objects.filter(id=expense_id).first()
                    if expense:
                        title = expense.title
                        expense.delete()
                        messages.success(request, f"Successfully deleted expense: {title}")
            else:
                messages.error(request, "Permission Denied. Only the CEO can delete expense records.")
            return redirect('account_expenses')
            
    # Calculate metrics
    import nepali_datetime
    import datetime
    today_bs = nepali_datetime.date.today()
    week_start_bs = today_bs - datetime.timedelta(days=today_bs.weekday())

    daily_expenses = ExpenseRecord.objects.filter(date=str(today_bs)).aggregate(Sum('amount'))['amount__sum'] or 0
    weekly_expenses = ExpenseRecord.objects.filter(date__gte=str(week_start_bs)).aggregate(Sum('amount'))['amount__sum'] or 0
    total_expenses = ExpenseRecord.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Highest expense area
    category_totals = ExpenseRecord.objects.values('category').annotate(total=Sum('amount')).order_by('-total')
    highest_expense_category = category_totals.first() if category_totals else None
    
    # All records
    records = ExpenseRecord.objects.all().order_by('-date', '-created_at')
    
    # Generate CA Insights based on highest category
    ca_insight_title = "Overall Financial Health"
    ca_insight_detail = "Your spending is well diversified. Consider setting monthly caps on variable expenses like Marketing and Misc."
    
    if highest_expense_category:
        highest_cat = highest_expense_category['category']
        if highest_cat == 'Payroll':
            ca_insight_title = "Optimize Payroll Structure"
            ca_insight_detail = "Payroll is currently your highest expense. As your CA, I advise cross-training employees to maximize productivity without hiring more staff, and evaluating performance-based bonuses over fixed salary increments."
        elif highest_cat == 'Marketing':
            ca_insight_title = "Audit Marketing ROI"
            ca_insight_detail = "Your marketing spend is disproportionately high. Calculate Customer Acquisition Cost (CAC) and eliminate underperforming channels immediately. Focus on organic SEO and high-ROI targeted ads."
        elif highest_cat == 'Operations':
            ca_insight_title = "Streamline Operations"
            ca_insight_detail = "High operational costs indicate inefficiencies. Bulk purchase office supplies, renegotiate lease terms if possible, and implement strict approval workflows for daily operational spends."
        elif highest_cat == 'Procurement':
            ca_insight_title = "Renegotiate Vendor Contracts"
            ca_insight_detail = "Hardware procurement is eating your margins. Leverage bulk buying and request net-30 or net-60 payment terms from suppliers to improve cash flow."
        elif highest_cat == 'Utilities':
            ca_insight_title = "Energy & Resource Audit"
            ca_insight_detail = "Utility expenses are spiking. Ensure systems are powered down during non-business hours and negotiate internet/telecom packages for corporate bulk rates."
            
    context = {
        'daily_expenses': daily_expenses,
        'weekly_expenses': weekly_expenses,
        'total_expenses': total_expenses,
        'highest_expense_category': highest_expense_category,
        'records': records,
        'ca_insight_title': ca_insight_title,
        'ca_insight_detail': ca_insight_detail,
        'categories': ExpenseRecord.CATEGORY_CHOICES,
    }
    return render(request, 'expenses.html', context)

def export_expenses_excel(request):
    if 'logged_in_uid' not in request.session:
        return redirect('login_view')

    uid = request.session['logged_in_uid']
    system_user = SystemUserProfile.objects.filter(uid=uid).first()
    
    if system_user and not system_user.can_access_account_expenses:
        from django.contrib import messages
        messages.error(request, "Access Denied: You do not have permission to download Expenses.")
        return redirect('dashboard')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    from .models import ExpenseRecord
    import openpyxl
    from django.http import HttpResponse

    records = ExpenseRecord.objects.all().order_by('-date')
    if start_date and end_date:
        records = records.filter(date__range=[start_date, end_date])
    elif start_date:
        records = records.filter(date__gte=start_date)
    elif end_date:
        records = records.filter(date__lte=end_date)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Expense Records'

    worksheet.append(['Date', 'Title', 'Category', 'Amount', 'Description', 'Logged By'])

    for record in records:
        logged_by_name = record.logged_by.full_name if record.logged_by else 'Unknown'
        worksheet.append([
            str(record.date),
            record.title,
            record.category,
            float(record.amount),
            record.description or '',
            logged_by_name
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="expense_records.xlsx"'
    workbook.save(response)
    return response
